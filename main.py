import openpyxl

inv_file = openpyxl.load_workbook("Inventory.xlsx")
product_list = inv_file["Sheet1"]

products_per_supplier = {}
total_value_per_supplier = {}
product_under_25_inv = {}

for product_row in range(2,
                         product_list.max_row + 1):  # range(2, max_row + 1)  is [2,3, 4, ..., last row] (row 1 is columns title and don't have to be treated and max_row is total row -1)
    supplier_name = product_list.cell(product_row, 4).value  # each cell for the column 4 Suppliers
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_number = product_list.cell(product_row, 1).value

    # calculate how products per supplier
    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = current_num_products + 1
    else:
        print("adding a new supplier")
        products_per_supplier[supplier_name] = 1

    # calculate total value of inventory per supplier
    if supplier_name in total_value_per_supplier :
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # calculate logic product inventory less than 25
    if inventory < 25:
        product_under_25_inv[product_number] = int(inventory)

print(product_under_25_inv)
